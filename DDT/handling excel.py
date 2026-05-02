import openpyxl

# writing data into excel

path=r'C:\Users\MD. MIDELAJ\Desktop\SELENIUM\DDT\DDT.xlsx'
workbook=openpyxl.load_workbook(path)
sheet=workbook["Sheet1"]
for i in range(1,7):
    for j in range(1,4):
        sheet.cell(i,j).value=input("enter the data  ")
workbook.save(path)
# reading data into excel

path=r'C:\Users\MD. MIDELAJ\Desktop\SELENIUM\DDT\DDT.xlsx'
workbook=openpyxl.load_workbook(path)
sheet=workbook["Sheet1"]
rows=sheet.max_row
cols=sheet.max_column
for i in range(1,rows+1):
    for j in range(1,cols+1):
        print(sheet.cell(i,j).value,end=" ")
    print()

